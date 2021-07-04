from bs4 import *
import openpyxl
import requests
from openpyxl import Workbook

#xlsx Initalization
wb = openpyxl.load_workbook("/home/mojo/Desktop/scrape/test.xlsx")
ws = wb.active

for row in range(2,ws.max_row+1):
    if ws.cell(row=row,column=6).value != None:
        print('B1 - If Loop')
        try:
            print('C1 - first Try')
            link = ws.cell(row=row,column=6).value
            headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:50.0) Gecko/20100101 Firefox/50.0'}
            req = requests.get(link, headers=headers)
            link = str(link)
            linkList = link.split("/") #URL Split Array
            proSkuCode = linkList[4]
            soup = BeautifulSoup(req.content,'html.parser')
            print('Valid Link')
        except :
            print('C1 - Fail')
            invalidLink = True
            print('Invalid Link')
        try:
            try:
                productPrice = soup.find("span",{"id":"priceblock_dealprice"}).text
            except :
                productPrice = soup.find("span",{"id":"priceblock_ourprice"}).text
        except:
            None
        try:
            proMrp = str(soup.find("span",{"class":"priceBlockStrikePriceString"}).text.strip())
            ws.cell(row=row,column=4,value=str(proMrp))
        except:
            proMrp = productPrice
            ws.cell(row=row,column=4,value=str(proMrp))

        if ws.cell(row=row,column=2).value != productPrice:
            ws.cell(row=row,column=3,value=str(productPrice))
        else :
            None
        if ws.cell(row=row,column=4).value != proMrp:
            ws.cell(row=row,column=5,value=str(proMrp))
        else :
            None
        wb.save("/home/mojo/Desktop/scrape/test.xlsx")
    else:
        break
